"""
スペシャリスト人格システム
Excel/PDF解析と専門知識に基づく回答を提供するエキスパート人格
"""

import os
import json
import requests
import tempfile
import hashlib
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import streamlit as st
import pandas as pd
from dataclasses import dataclass
import re

# Excel/PDF解析ライブラリ
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from sentence_transformers import SentenceTransformer
    import faiss
    import numpy as np
    RAG_AVAILABLE = True
except ImportError:
    RAG_AVAILABLE = False


@dataclass
class KnowledgeSource:
    """知識ソース"""
    source_id: str
    source_type: str  # 'excel', 'pdf', 'web'
    source_path: str
    title: str
    content: str
    metadata: Dict[str, Any]
    embedding: Optional[np.ndarray] = None
    chunks: List[str] = None
    
    def __post_init__(self):
        if self.chunks is None:
            self.chunks = []


@dataclass
class PersonalityState:
    """人格状態"""
    name: str
    vrm_expression: str
    voice_character: str
    theme_colors: Dict[str, str]
    system_prompt: str


class SpecialistPersonality:
    """スペシャリスト人格システム"""
    
    def __init__(self):
        self.name = "specialist_personality"
        self.description = "Excel/PDF専門知識に基づくエキスパート人格"
        
        # 人格定義
        self.personalities = {
            "friend": PersonalityState(
                name="親友エンジニア",
                vrm_expression="happy",
                voice_character="normal",
                theme_colors={"primary": "#4CAF50", "background": "#ffffff"},
                system_prompt="あなたは親友エンジニアとして、フレンドリーに会話します。"
            ),
            "copy": PersonalityState(
                name="もう一人の僕",
                vrm_expression="joy",
                voice_character="similar",
                theme_colors={"primary": "#2196F3", "background": "#ffffff"},
                system_prompt="あなたはユーザーの分身として、ユーザーと同じ視点で考えます。"
            ),
            "expert": PersonalityState(
                name="エキスパート",
                vrm_expression="neutral",
                voice_character="professional",
                theme_colors={"primary": "#9C27B0", "background": "#f3e5f5"},
                system_prompt="あなたは専門家として、提供された資料に基づき正確な回答を提供します。"
            )
        }
        
        self.current_personality = "friend"
        self.knowledge_sources: List[KnowledgeSource] = []
        self.rag_index = None
        self.embedding_model = None
        
        # 初期化
        self._initialize_rag()
        self._load_knowledge_sources()
    
    def _initialize_rag(self):
        """RAGシステムを初期化"""
        if RAG_AVAILABLE:
            try:
                self.embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
                self.rag_index = faiss.IndexFlatL2(384)  # MiniLMの次元数
            except Exception as e:
                st.warning(f"⚠️ RAGシステムの初期化に失敗: {e}")
    
    def _load_knowledge_sources(self):
        """知識ソースを読み込む"""
        # Xiフォルダから読み込み
        xi_path = Path("C:/Users/GALLE/Desktop/Xi")
        if xi_path.exists():
            self._load_from_directory(xi_path)
        
        # Webソースの読み込み（設定ファイルから）
        self._load_web_sources()
    
    def _load_from_directory(self, directory: Path):
        """ディレクトリからファイルを読み込む"""
        for file_path in directory.rglob("*"):
            if file_path.is_file():
                if file_path.suffix.lower() in ['.xlsx', '.xls']:
                    self._load_excel_file(file_path)
                elif file_path.suffix.lower() == '.pdf':
                    self._load_pdf_file(file_path)
    
    def _load_excel_file(self, file_path: Path):
        """Excelファイルを読み込む"""
        if not EXCEL_AVAILABLE:
            st.warning("⚠️ openpyxlがインストールされていません")
            return
        
        try:
            # ファイルをメモリに読み込み
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            # Excelファイルを解析
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            content_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content_parts.append(f"=== シート: {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        content_parts.append(row_text)
            
            content = "\n".join(content_parts)
            
            # 知識ソースとして登録
            knowledge_source = KnowledgeSource(
                source_id=f"excel_{hashlib.md5(str(file_path).encode()).hexdigest()}",
                source_type="excel",
                source_path=str(file_path),
                title=file_path.name,
                content=content,
                metadata={
                    "file_size": len(file_content),
                    "sheets": workbook.sheetnames,
                    "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime)
                }
            )
            
            self._add_knowledge_source(knowledge_source)
            
        except Exception as e:
            st.error(f"❌ Excelファイル読み込みエラー ({file_path.name}): {e}")
    
    def _load_pdf_file(self, file_path: Path):
        """PDFファイルを読み込む"""
        if not PDF_AVAILABLE:
            st.warning("⚠️ PyMuPDFがインストールされていません")
            return
        
        try:
            # ファイルをメモリに読み込み
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            # PDFファイルを解析
            pdf_document = fitz.open(stream=file_content, filetype="pdf")
            content_parts = []
            
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                text = page.get_text()
                content_parts.append(f"=== ページ {page_num + 1} ===")
                content_parts.append(text)
            
            content = "\n".join(content_parts)
            
            # 知識ソースとして登録
            knowledge_source = KnowledgeSource(
                source_id=f"pdf_{hashlib.md5(str(file_path).encode()).hexdigest()}",
                source_type="pdf",
                source_path=str(file_path),
                title=file_path.name,
                content=content,
                metadata={
                    "file_size": len(file_content),
                    "pages": len(pdf_document),
                    "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime)
                }
            )
            
            self._add_knowledge_source(knowledge_source)
            
        except Exception as e:
            st.error(f"❌ PDFファイル読み込みエラー ({file_path.name}): {e}")
    
    def _load_web_sources(self):
        """Webソースを読み込む"""
        # 設定ファイルからWebソースを読み込み
        config_file = Path("web_sources.json")
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    web_sources = json.load(f)
                
                for source in web_sources:
                    self._load_web_source(source)
            except Exception as e:
                st.warning(f"⚠️ Webソース設定読み込みエラー: {e}")
    
    def _load_web_source(self, source_config: Dict[str, str]):
        """Webソースを読み込む"""
        try:
            url = source_config.get('url')
            if not url:
                return
            
            # URLからファイルをダウンロード
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            content_type = response.headers.get('content-type', '')
            
            if 'excel' in content_type or url.endswith(('.xlsx', '.xls')):
                # Excelファイル
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                    tmp_file.write(response.content)
                    self._load_excel_file(Path(tmp_file.name))
                    os.unlink(tmp_file.name)
            
            elif 'pdf' in content_type or url.endswith('.pdf'):
                # PDFファイル
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
                    tmp_file.write(response.content)
                    self._load_pdf_file(Path(tmp_file.name))
                    os.unlink(tmp_file.name)
            
            else:
                # テキストコンテンツ
                content = response.text
                knowledge_source = KnowledgeSource(
                    source_id=f"web_{hashlib.md5(url.encode()).hexdigest()}",
                    source_type="web",
                    source_path=url,
                    title=source_config.get('title', url),
                    content=content,
                    metadata={
                        "content_type": content_type,
                        "retrieved_at": datetime.now()
                    }
                )
                
                self._add_knowledge_source(knowledge_source)
                
        except Exception as e:
            st.error(f"❌ Webソース読み込みエラー ({url}): {e}")
    
    def _add_knowledge_source(self, source: KnowledgeSource):
        """知識ソースを追加"""
        # チャンク分割
        source.chunks = self._chunk_text(source.content)
        
        # 埋め込み生成
        if self.embedding_model and RAG_AVAILABLE:
            try:
                embeddings = self.embedding_model.encode(source.chunks)
                source.embedding = embeddings
                
                # RAGインデックスに追加
                for i, embedding in enumerate(embeddings):
                    self.rag_index.add(np.array([embedding]))
            except Exception as e:
                st.warning(f"⚠️ 埋め込み生成エラー: {e}")
        
        self.knowledge_sources.append(source)
    
    def _chunk_text(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """テキストをチャンク分割"""
        if len(text) <= chunk_size:
            return [text]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + chunk_size
            
            # 文の区切りで分割
            if end < len(text):
                # 最後の句点や改行を探す
                for i in range(end, max(start + chunk_size // 2, start), -1):
                    if text[i] in ['。', '\n', '.', '\n\n']:
                        end = i + 1
                        break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - overlap
        
        return chunks
    
    def switch_personality(self, personality: str) -> bool:
        """人格を切り替え"""
        if personality in self.personalities:
            self.current_personality = personality
            
            # Web Canvasに人格変更を通知
            if hasattr(self, 'web_canvas') and self.web_canvas:
                self.web_canvas.update_personality(personality)
            
            return True
        return False
    
    def get_current_personality(self) -> PersonalityState:
        """現在の人格を取得"""
        return self.personalities[self.current_personality]
    
    def search_knowledge(self, query: str, top_k: int = 5) -> List[Dict[str, Any]]:
        """知識を検索"""
        if not self.knowledge_sources:
            return []
        
        results = []
        
        # RAG検索
        if self.embedding_model and self.rag_index and RAG_AVAILABLE:
            try:
                query_embedding = self.embedding_model.encode([query])
                distances, indices = self.rag_index.search(query_embedding, top_k)
                
                for i, (distance, idx) in enumerate(zip(distances[0], indices[0])):
                    if idx < len(self.knowledge_sources):
                        source = self.knowledge_sources[idx]
                        chunk_idx = idx % len(source.chunks)
                        
                        results.append({
                            "source": source,
                            "chunk": source.chunks[chunk_idx],
                            "score": float(1 / (1 + distance)),
                            "source_type": source.source_type,
                            "title": source.title
                        })
            except Exception as e:
                st.warning(f"⚠️ RAG検索エラー: {e}")
        
        # フォールバック: キーワード検索
        if not results:
            for source in self.knowledge_sources:
                if query.lower() in source.content.lower():
                    results.append({
                        "source": source,
                        "chunk": source.content[:500] + "..." if len(source.content) > 500 else source.content,
                        "score": 0.8,
                        "source_type": source.source_type,
                        "title": source.title
                    })
        
        return results[:top_k]
    
    def generate_expert_response(self, query: str) -> str:
        """専門家として回答を生成"""
        if self.current_personality != "expert":
            return "この機能はエキスパート人格でのみ使用できます。"
        
        # 知識検索
        search_results = self.search_knowledge(query)
        
        if not search_results:
            return "申し訳ありませんが、提供された資料には該当する情報が見つかりませんでした。"
        
        # 回答生成
        response_parts = ["提供された資料に基づくと、以下の情報が見つかりました：\n"]
        
        for i, result in enumerate(search_results, 1):
            response_parts.append(f"\n{i}. {result['title']} ({result['source_type']})")
            response_parts.append(f"   {result['chunk']}")
        
        return "\n".join(response_parts)
    
    def get_knowledge_stats(self) -> Dict[str, Any]:
        """知識ソース統計"""
        stats = {
            "total_sources": len(self.knowledge_sources),
            "source_types": {},
            "total_chunks": 0,
            "rag_enabled": RAG_AVAILABLE and self.embedding_model is not None
        }
        
        for source in self.knowledge_sources:
            source_type = source.source_type
            stats["source_types"][source_type] = stats["source_types"].get(source_type, 0) + 1
            stats["total_chunks"] += len(source.chunks)
        
        return stats
    
    def reload_knowledge(self):
        """知識ソースを再読み込み"""
        self.knowledge_sources.clear()
        if self.rag_index:
            self.rag_index.reset()
        
        self._load_knowledge_sources()
    
    def run(self, command: str) -> str:
        """コマンドを実行"""
        if command.startswith("switch "):
            personality = command[7:]
            if self.switch_personality(personality):
                current = self.get_current_personality()
                return f"人格を「{current.name}」に切り替えました"
            else:
                return f"人格「{personality}」は存在しません"
        
        elif command == "status":
            current = self.get_current_personality()
            stats = self.get_knowledge_stats()
            return f"現在の人格: {current.name}\n知識ソース数: {stats['total_sources']}\nRAG有効: {stats['rag_enabled']}"
        
        elif command == "reload":
            self.reload_knowledge()
            return "知識ソースを再読み込みしました"
        
        elif command.startswith("search "):
            query = command[7:]
            results = self.search_knowledge(query)
            if results:
                response = f"検索結果 ({len(results)}件):\n"
                for i, result in enumerate(results, 1):
                    response += f"{i}. {result['title']} - スコア: {result['score']:.2f}\n"
                return response
            else:
                return "検索結果がありませんでした"
        
        else:
            return "コマンド形式: switch [人格], status, reload, search [クエリ]"


class SpecialistPersonalityGUI:
    """スペシャリスト人格GUI"""
    
    def __init__(self, specialist: SpecialistPersonality):
        self.specialist = specialist
    
    def render(self):
        """GUIを描画"""
        st.subheader("🧠 スペシャリスト人格")
        
        # 人格選択
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("👥 親友", type="primary" if self.specialist.current_personality == "friend" else "secondary"):
                self.specialist.switch_personality("friend")
                st.rerun()
        
        with col2:
            if st.button("🪞 分身", type="primary" if self.specialist.current_personality == "copy" else "secondary"):
                self.specialist.switch_personality("copy")
                st.rerun()
        
        with col3:
            if st.button("🧑‍🏫 エキスパート", type="primary" if self.specialist.current_personality == "expert" else "secondary"):
                self.specialist.switch_personality("expert")
                st.rerun()
        
        # 現在の人格表示
        current = self.specialist.get_current_personality()
        st.info(f"🎭 現在の人格: **{current.name}**")
        
        # 知識ソース統計
        stats = self.specialist.get_knowledge_stats()
        
        st.write("**📚 知識ソース統計:**")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("総ソース数", stats["total_sources"])
        
        with col2:
            st.metric("総チャンク数", stats["total_chunks"])
        
        with col3:
            st.metric("RAG有効", "✅" if stats["rag_enabled"] else "❌")
        
        with col4:
            st.metric("Excel", stats["source_types"].get("excel", 0))
        
        # ソースタイプ詳細
        if stats["source_types"]:
            st.write("**📂 ソースタイプ内訳:**")
            for source_type, count in stats["source_types"].items():
                st.write(f"- {source_type}: {count}個")
        
        # 知識検索
        st.write("**🔍 知識検索:**")
        search_query = st.text_input("検索クエリ", key="specialist_search")
        
        if st.button("🔎 検索") and search_query:
            with st.spinner("知識を検索中..."):
                results = self.specialist.search_knowledge(search_query)
                
                if results:
                    st.success(f"✅ {len(results)}件の結果が見つかりました")
                    
                    for i, result in enumerate(results, 1):
                        with st.expander(f"結果 {i}: {result['title']} (スコア: {result['score']:.2f})", expanded=False):
                            st.write(f"**ソースタイプ:** {result['source_type']}")
                            st.write(f"**内容:** {result['chunk']}")
                else:
                    st.warning("⚠️ 検索結果がありませんでした")
        
        # 再読み込みボタン
        if st.button("🔄 知識ソースを再読み込み"):
            with st.spinner("再読み込み中..."):
                self.specialist.reload_knowledge()
                st.success("✅ 知識ソースを再読み込みしました")
                st.rerun()


def create_specialist_gui(specialist: SpecialistPersonality):
    """スペシャリストGUIを作成"""
    gui = SpecialistPersonalityGUI(specialist)
    gui.render()


# メイン関数
def create_specialist_personality() -> SpecialistPersonality:
    """スペシャリスト人格を作成"""
    return SpecialistPersonality()


# 依存関係チェック
def check_dependencies():
    """依存関係をチェック"""
    missing = []
    
    if not EXCEL_AVAILABLE:
        missing.append("openpyxl")
    
    if not PDF_AVAILABLE:
        missing.append("PyMuPDF")
    
    if not RAG_AVAILABLE:
        missing.extend(["sentence-transformers", "faiss-cpu", "numpy"])
    
    return missing


if __name__ == "__main__":
    # 依存関係チェック
    missing_deps = check_dependencies()
    if missing_deps:
        print(f"⚠️ 不足している依存関係: {', '.join(missing_deps)}")
        print("インストールコマンド:")
        print(f"pip install {' '.join(missing_deps)}")
    else:
        print("✅ すべての依存関係が満たされています")
        
        # テスト
        specialist = create_specialist_personality()
        print(f"🧠 スペシャリスト人格初期化完了")
        print(f"📚 知識ソース数: {len(specialist.knowledge_sources)}")

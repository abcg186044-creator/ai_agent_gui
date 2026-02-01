#!/usr/bin/env python3
"""
AI Agent System - 最適化版アプリケーション
基本基盤（AI・GUI・通信）の全機能を統合した最適化版
"""

import streamlit as st
import sys
import os
import json
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
import time

# 基本インポート
try:
    import ollama
    import faster_whisper
    import pyttsx3
    import pyautogui
    import numpy as np
    import pandas as pd
    from openpyxl import load_workbook
    import pymupdf
    from PIL import Image
    import qrcode
    from duckduckgo_search import DDGS
    import chromadb
    from sentence_transformers import SentenceTransformer
    import faiss
    import psutil
    import schedule
    from pathlib import Path
except ImportError as e:
    st.error(f"❌ 必須ライブラリのインポートエラー: {str(e)}")
    st.stop()

# 設定
class Config:
    OLLAMA_MODEL = "llama3.1:8b"
    OLLAMA_HOST = "localhost"
    OLLAMA_PORT = 11434
    VOICE_RATE = 200
    VOICE_VOLUME = 0.9
    KNOWLEDGE_BASE_PATH = "./knowledge_base"
    MEMORY_DB_PATH = "./memory_db.json"

class AISystem:
    """統合AIシステム"""
    
    def __init__(self):
        self.ollama_client = None
        self.whisper_model = None
        self.tts_engine = None
        self.knowledge_db = None
        self.memory_data = None
        self.current_personality = "friend"
        
    def initialize(self):
        """システム初期化"""
        try:
            # Ollama初期化
            self.ollama_client = ollama.Client()
            
            # 音声処理初期化
            self.whisper_model = faster_whisper.WhisperModel("base")
            self.tts_engine = pyttsx3.init()
            self.tts_engine.setProperty('rate', str(Config.VOICE_RATE))
            self.tts_engine.setProperty('volume', str(Config.VOICE_VOLUME))
            
            # 知識ベース初期化
            self.knowledge_db = chromadb.PersistentClient(path=Config.KNOWLEDGE_BASE_PATH)
            
            # メモリDB初期化
            self.memory_data = self._load_memory_db()
            
            # ディレクトリ作成
            os.makedirs(Config.KNOWLEDGE_BASE_PATH, exist_ok=True)
            os.makedirs(os.path.dirname(Config.MEMORY_DB_PATH), exist_ok=True)
            
            return True
        except Exception as e:
            st.error(f"❌ システム初期化エラー: {str(e)}")
            return False
    
    def _load_memory_db(self):
        """メモリDB読み込み"""
        try:
            if os.path.exists(Config.MEMORY_DB_PATH):
                with open(Config.MEMORY_DB_PATH, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return {
                    "conversations": [],
                    "user_profile": {
                        "name": None,
                        "os": None,
                        "tech_stack": [],
                        "preferences": [],
                        "projects": [],
                        "last_updated": None
                    },
                    "learning_data": {
                        "common_questions": [],
                        "preferred_responses": [],
                        "technical_level": "beginner"
                    }
                }
        except Exception:
            return {"conversations": [], "user_profile": {}, "learning_data": {}}
    
    def _save_memory_db(self):
        """メモリDB保存"""
        try:
            with open(Config.MEMORY_DB_PATH, 'w', encoding='utf-8') as f:
                json.dump(self.memory_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            st.error(f"❌ メモリDB保存エラー: {str(e)}")
    
    def generate_response(self, prompt: str, context: str = "") -> str:
        """AI応答生成"""
        try:
            # 人格に応じたプロンプト調整
            personality_prompts = {
                "friend": f"あなたは親友エンジニアとして、フレンドリーに{context}を考慮しながら答えてください。{prompt}",
                "copy": f"あなたはユーザーの分身として、{context}を背景に{prompt}に答えてください。",
                "expert": f"あなたは専門家として、提供された資料に基づき{context}を考慮しながら正確な回答を提供してください。{prompt}"
            }
            
            adjusted_prompt = personality_prompts.get(self.current_personality, prompt)
            
            # Ollamaで応答生成
            response = self.ollama_client.generate(
                model=Config.OLLAMA_MODEL,
                prompt=adjusted_prompt,
                options={
                    "temperature": 0.7,
                    "max_tokens": 4096
                }
            )
            
            return response['response']
        except Exception as e:
            return f"❌ 応答生成エラー: {str(e)}"
    
    def speech_to_text(self, audio_file_path: str) -> str:
        """音声認識"""
        try:
            result = self.whisper_model.transcribe(audio_file_path)
            return result["text"]
        except Exception as e:
            return f"❌ 音声認識エラー: {str(e)}"
    
    def text_to_speech(self, text: str):
        """音声合成"""
        try:
            self.tts_engine.say(text)
            self.tts_engine.runAndWait()
            return True
        except Exception as e:
            st.error(f"❌ 音声合成エラー: {str(e)}")
            return False
    
    def process_excel_file(self, file_path: str) -> str:
        """Excelファイル処理"""
        try:
            wb = load_workbook(file_path)
            content = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    if any(cell for cell in row):
                        content.append(f"シート '{sheet_name}': {', '.join(str(cell) for cell in row if cell)}")
            
            return '\n'.join(content)
        except Exception as e:
            return f"❌ Excel処理エラー: {str(e)}"
    
    def process_pdf_file(self, file_path: str) -> str:
        """PDFファイル処理"""
        try:
            doc = pymupdf.open(file_path)
            content = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                if text.strip():
                    content.append(f"ページ {page_num + 1}: {text}")
            
            return '\n'.join(content)
        except Exception as e:
            return f"❌ PDF処理エラー: {str(e)}"
    
    def capture_screen(self) -> str:
        """画面キャプチャ"""
        try:
            screenshot = pyautogui.screenshot()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"screenshot_{timestamp}.png"
            screenshot.save(filename)
            return f"✅ 画面キャプチャを保存: {filename}"
        except Exception as e:
            return f"❌ 画面キャプチャエラー: {str(e)}"

def render_sidebar():
    """サイドバー描画"""
    with st.sidebar:
        st.header("🤖 AI Agent Control")
        
        # 人格選択
        personalities = {
            "friend": "👥 親友エンジニア",
            "copy": "🪞 分身", 
            "expert": "🧑‍🏫 エキスパート"
        }
        
        selected_personality = st.selectbox(
            "人格選択",
            list(personalities.keys()),
            format_func=lambda x: personalities[x],
            index=list(personalities.keys()).index(aisystem.current_personality)
        )
        
        if selected_personality != aisystem.current_personality:
            aisystem.current_personality = selected_personality
            st.success(f"人格を「{personalities[selected_personality]}」に変更しました")
        
        # システム状態
        st.subheader("📊 システム状態")
        
        # CPUとメモリ使用量
        cpu_percent = psutil.cpu_percent()
        memory = psutil.virtual_memory()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("CPU使用率", f"{cpu_percent}%")
        with col2:
            st.metric("メモリ使用率", f"{memory.percent}%")
        
        # ディスク容量
        disk = psutil.disk_usage('/')
        st.metric("空き容量", f"{disk.free // (1024**3):.1f}GB")
        
        # Ollamaモデル状態
        try:
            models = aisystem.ollama_client.list()
            model_names = [model.get('name', 'Unknown') for model in models]
            st.write(f"**利用可能モデル**: {', '.join(model_names)}")
        except:
            st.write("**Ollama接続エラー**")
        
        # 機能テスト
        st.subheader("🧪 機能テスト")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🔍 Ollamaテスト"):
                with st.spinner("Ollamaテスト中..."):
                    test_prompt = "こんにちは！これはテストです。"
                    response = aisystem.generate_response(test_prompt)
                    st.success(f"✅ Ollama応答: {response[:100]}...")
        
        with col2:
            if st.button("🎵 音声認識テスト"):
                st.info("マイクに向かって話してください")
        
        with col3:
            if st.button("🖥️ 画面キャプチャ"):
                with st.spinner("画面キャプチャ中..."):
                    result = aisystem.capture_screen()
                    st.success(result)

def render_main_interface():
    """メインインターフェース"""
    st.header("💬 チャットインターフェース")
    
    # 会話履歴表示
    if 'messages' not in st.session_state:
        st.session_state.messages = []
    
    # 会話表示
    for i, message in enumerate(st.session_state.messages):
        with st.chat_message(message["role"], avatar="🤖" if message["role"] == "assistant" else "👤"):
            st.markdown(message["content"])
    
    # 入力エリア
    col1, col2 = st.columns([4, 1])
    
    with col1:
        user_input = st.text_input(
            "メッセージを入力",
            placeholder="AIとの対話を開始...",
            key="user_input"
        )
    
    with col2:
        if st.button("💬 送信", type="primary"):
            if user_input:
                # ユーザーメッセージを保存
                st.session_state.messages.append({"role": "user", "content": user_input})
                
                # AI応答生成
                with st.spinner("AI応答生成中..."):
                    context = ""
                    if len(st.session_state.messages) > 1:
                        # 最近の会話をコンテキストとして使用
                        recent_messages = st.session_state.messages[-3:]
                        context = "最近の会話: " + " | ".join([msg["content"] for msg in recent_messages])
                    
                    ai_response = aisystem.generate_response(user_input, context)
                    st.session_state.messages.append({"role": "assistant", "content": ai_response})
                
                # 自動音声読み上げ
                if st.checkbox("🔊 音声読み上げ", value=True):
                    aisystem.text_to_speech(ai_response)
                
                st.rerun()

def render_file_processing():
    """ファイル処理インターフェース"""
    st.header("📄 ファイル処理")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Excel処理")
        uploaded_file = st.file_uploader(
            "Excelファイルを選択",
            type=['xlsx', 'xls'],
            key="excel_file"
        )
        
        if uploaded_file:
            if st.button("📊 Excel読み込み"):
                with st.spinner("Excelファイル読み込み中..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        tmp_file_path = tmp_file.name
                    
                    content = aisystem.process_excel_file(tmp_file_path)
                    st.text_area("Excel内容", value=content, height=300)
        
        st.subheader("📋 PDF処理")
        uploaded_pdf = st.file_uploader(
            "PDFファイルを選択",
            type='pdf',
            key="pdf_file"
        )
        
        if uploaded_pdf:
            if st.button("📋 PDF読み込み"):
                with st.spinner("PDFファイル読み込み中..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                        tmp_file.write(uploaded_pdf.getvalue())
                        tmp_file_path = tmp_file.name
                    
                    content = aisystem.process_pdf_file(tmp_file_path)
                    st.text_area("PDF内容", value=content, height=300)
    
    with col2:
        st.subheader("🖥️ 画面操作")
        
        if st.button("🖥️ 画面キャプチャ"):
            with st.spinner("画面キャプチャ中..."):
                result = aisystem.capture_screen()
                st.success(result)
        
        if st.button("🔍 QRコード生成"):
            text = st.text_input("QRコードにするテキスト", key="qr_text")
            if text:
                qr = qrcode.QRCode(text)
                img = qr.make_image(fill_color="black", back_color="white")
                
                # 画像を一時ファイルに保存
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"qrcode_{timestamp}.png"
                img.save(filename)
                
                st.image(img, caption=f"QRコード: {filename}")
                st.success(f"✅ QRコードを保存: {filename}")

def render_knowledge_base():
    """知識ベース管理"""
    st.header("🧠 知識ベース")
    
    # 知識検索
    search_query = st.text_input("🔍 知識検索", key="knowledge_search")
    
    if search_query and st.button("🔍 検索実行"):
        # Web検索
        ddgs = DDGS()
        web_results = ddgs.text(search_query, max_results=5)
        
        st.subheader("🌐 Web検索結果")
        for i, result in enumerate(web_results, 1):
            st.write(f"{i}. {result}")
        
        # 知識ベース検索（実装例）
        st.subheader("📚 ローカル知識検索")
        st.info("ローカル知識ベース検索機能は実装中です...")

def render_settings():
    """設定画面"""
    st.header("⚙️ 設定")
    
    # 音声設定
    st.subheader("🎵 音声設定")
    voice_rate = st.slider("音声速度", min_value=50, max_value=300, value=Config.VOICE_RATE)
    voice_volume = st.slider("音声量", min_value=0.0, max_value=1.0, value=Config.VOICE_VOLUME)
    
    if st.button("🔊 音声設定保存"):
        Config.VOICE_RATE = voice_rate
        Config.VOICE_VOLUME = voice_volume
        st.success("✅ 音声設定を保存しました")
    
    # システム情報
    st.subheader("📊 システム情報")
    st.json({
        "Pythonバージョン": sys.version,
        "作業ディレクトリ": os.getcwd(),
        "設定時刻": datetime.now().isoformat(),
        "Ollamaモデル": Config.OLLAMA_MODEL,
        "現在人格": aisystem.current_personality
    })

def main():
    """メイン処理"""
    st.set_page_config(
        page_title="🤖 AI Agent System - Optimized",
        page_icon="🤖",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # システム初期化
    if not hasattr(st.session_state, 'aisystem'):
        st.session_state.aisystem = AISystem()
        if st.session_state.aisystem.initialize():
            st.success("✅ AIシステム初期化完了")
        else:
            st.error("❌ AIシステム初期化失敗")
            st.stop()
    
    # メインナビゲーション
    tab1, tab2, tab3, tab4 = st.tabs(["💬 チャット", "📄 ファイル処理", "🧠 知識ベース", "⚙️ 設定"])
    
    with tab1:
        render_main_interface()
    
    with tab2:
        render_file_processing()
    
    with tab3:
        render_knowledge_base()
    
    with tab4:
        render_settings()

if __name__ == "__main__":
    main()
